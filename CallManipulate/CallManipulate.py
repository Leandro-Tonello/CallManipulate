import re
import openpyxl
import pandas as pd
import datetime
import random
import string
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from tkinter import ttk


# Definicion de la clase Persona
class Persona:
    def __init__(self, numero_cliente, nombre, telefono, documento, tarjetas):
        self.numero_cliente = numero_cliente
        self.nombre = nombre
        self.telefono = telefono
        self.documento = documento
        self.tarjetas = self.etiquetar_tarjetas(tarjetas)

    def etiquetar_tarjetas(self, tarjetas):
        etiquetas = []
        for tarjeta in tarjetas:
            if tarjeta.startswith("4"):
                etiquetas.append("Visa: " + tarjeta)
            elif tarjeta.startswith("5"):
                etiquetas.append("Master: " + tarjeta)
            elif tarjeta.startswith("3"):
                etiquetas.append("Amex: " + tarjeta)
            else:
                etiquetas.append(tarjeta)
        return etiquetas


def generar_archivo():
    try:
        # Obtener la ubicacion del archivo de entrada
        ruta_archivo = entrada_archivo.get()

        if not ruta_archivo:
            raise ValueError("No se selecciono un archivo de entrada.")

        # Busqueda de la info en el Excel
        datos = pd.read_excel(ruta_archivo, usecols=[0], engine="openpyxl")
        columna_0 = datos.iloc[:, 0]  # Acceder a la columna 0 del DataFrame
        texto = ""
        for valor in columna_0:
            texto += str(valor) + "\n"

        # Patrones de expresiones regulares
        patron_persona = r"\n(\d+)\n\s+Nombre\s+:\s+([^:\n]+)\s+Doc\.\s+:\s+DU\s+([^:\n]+)([\s\S]+?)(?=\n\d+\n|\Z)"
        patron_telefono = r"Telefono\s+:\s+([^:\n]+)"
        patron_tarjeta = r"\b\d{15,16}\b"

        # Extraccion de informacion de todas las personas
        personas = []
        matches_personas = re.finditer(patron_persona, texto, re.MULTILINE)
        for match_persona in matches_personas:
            numero_cliente = match_persona.group(1).strip()
            nombre = match_persona.group(2).strip()
            documento = match_persona.group(3).strip()
            bloque_persona = match_persona.group(4)

            telefono = re.search(patron_telefono, bloque_persona).group(1).strip().replace(" Int.", "").replace(">", "").replace("Int.", "")
            tarjetas = re.findall(patron_tarjeta, bloque_persona)

            persona = Persona(numero_cliente, nombre, telefono, documento, tarjetas)
            personas.append(persona)

        # Crear un libro y una hoja de calculo con openpyxl
        workbook = Workbook()
        worksheet = workbook.active

        # Crear el DataFrame
        data = []
        for persona in personas:
            tarjetas = ', '.join(persona.tarjetas)
            data.append([persona.numero_cliente, persona.nombre, persona.telefono, tarjetas])

        df = pd.DataFrame(data, columns=['Nro', 'Nombre', 'Telefono', 'Tarjetas'])

        # Escribir los datos en la hoja de calculo
        for row_num, row_data in enumerate(df.values, 2):
            for col_num, value in enumerate(row_data, 1):
                cell = worksheet.cell(row=row_num, column=col_num, value=value)

        # Configurar el formato y estilo
        header_fill = PatternFill(fill_type='solid', fgColor='000000')
        header_font = Font(color='FFFFFF', bold=True)
        data_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        gray_fill = PatternFill(fill_type='solid', fgColor='E9E9E9')

        # Aplicar formato y estilo a los encabezados
        for col_num, header in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Aplicar formato y estilo a las celdas de datos
        for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=df.shape[1]):
            for cell in row:
                cell.border = data_border

        # Aplicar formato y estilo a las filas alternas
        for row_num in range(2, worksheet.max_row + 1, 2):
            for col_num in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.fill = gray_fill

        # Ajustar el ancho de las columnas
        for column in worksheet.columns:
            max_length = max(len(str(cell.value)) for cell in column)
            adjusted_width = (max_length + 2) * 1.2
            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # Generar un nombre de archivo aleatorio basado en la fecha y hora
        timestamp = datetime.datetime.now().strftime("%Y%m%d")
        random_word = ''.join(random.choices(string.ascii_lowercase, k=5))
        nombre_archivo = f"{timestamp}_{random_word}.xlsx"

        # Ruta de guardado con el nombre de archivo generado
        ruta_guardado = filedialog.asksaveasfilename(defaultextension=".xlsx")

        if not ruta_guardado:
            raise ValueError("No se selecciono una ubicacion de guardado.")

        # Guardar el archivo Excel
        workbook.save(ruta_guardado)
        workbook.close()

        # Mostrar la ubicacion del archivo generado
        resultado_archivo.config(text="Archivo generado:\n" + ruta_guardado)
        messagebox.showinfo("Proceso completado", "El archivo se ha generado correctamente.")

    except Exception as e:
        messagebox.showerror("Error", str(e))

# Funci�n para seleccionar un archivo
def seleccionar_archivo():
    archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx")])
    entrada_archivo.delete(0, tk.END)
    entrada_archivo.insert(0, archivo)

# Crear la interfaz de usuario
window = tk.Tk()
window.title("Generador de Archivo Excel")

style = ttk.Style()
style.configure("TButton", font=("Arial", 12), padding=6)

# Etiqueta para mostrar el texto "Seleccione el Excel"
lbl_archivo = ttk.Label(window, text="Seleccione el Excel:", font=("Arial", 12, "bold"))
lbl_archivo.pack(pady=10)

# Marco que contiene el campo de entrada y el bot�n "Seleccionar"
frame_seleccionar = ttk.Frame(window)
frame_seleccionar.pack(pady=5)

# Campo de entrada deshabilitado para mostrar la ubicaci�n del archivo seleccionado
entrada_archivo = ttk.Entry(frame_seleccionar, width=50, font=("Arial", 12), state="enabled")
entrada_archivo.pack(side=tk.LEFT, padx=5)

# Bot�n para seleccionar el archivo
btn_seleccionar = ttk.Button(frame_seleccionar, text="Seleccionar", command=seleccionar_archivo)
btn_seleccionar.pack(side=tk.LEFT)

# Bot�n de ejecutar
btn_ejecutar = ttk.Button(window, text="Ejecutar", command=generar_archivo)
btn_ejecutar.pack(pady=10)

# Etiqueta para mostrar la ubicaci�n del archivo generado
resultado_archivo = ttk.Label(window, text="")
resultado_archivo.pack()

# Calcula el tama�o de la ventana teniendo en cuenta el espacio requerido por los elementos internos
window.update_idletasks()
width = max(window.winfo_reqwidth(), 800) + 10
height = max(window.winfo_reqheight(), 200) + 10
window.geometry(f"{width}x{height}")

window.mainloop()
